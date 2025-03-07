from flask import Flask, render_template, request, redirect, url_for, session, send_file
import sqlite3
import os
from functools import wraps
from fpdf import FPDF
from openpyxl import Workbook
from datetime import datetime
import pytz
import uuid

app = Flask(__name__)
app.secret_key = "senha_segura"  # Chave para gerenciar sessões

# Função para criar o banco de dados (simplificada)
def criar_banco():
    with sqlite3.connect('cadastros.db') as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS inscritos(
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      nome TEXT,
                      rg TEXT UNIQUE,
                      cpf TEXT UNIQUE,
                      celular TEXT,
                      email TEXT,
                      empresa TEXT,
                      bairro TEXT,
                      autorizacao_dados TEXT,
                      autorizacao_imagem TEXT
                      )''')

criar_banco()

# Função para verificar login
def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "logged_in" not in session:
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper

# Rotas
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form['usuario']
        senha = request.form['senha']
        if usuario == "SENALBA RIO CAPITAL" and senha == "Sen@lb#799":
            session["logged_in"] = True
            return redirect(url_for("inscritos"))
        else:
            return render_template("login.html", erro="Usuário ou senha incorretos!")
    return render_template("login.html")

@app.route('/logout')
def logout():
    session.pop("logged_in", None)
    return redirect(url_for("login"))

# Define o fuso horário
fuso_horario = pytz.timezone('America/Sao_Paulo')

# Data e hora limite com fuso correto
DATA_FINAL = fuso_horario.localize(datetime(2025, 4, 17, 17, 59))

@app.route('/')
def index():
    agora_naive = datetime.now() # Hora atual naive
    agora = fuso_horario.localize(agora_naive) # Hora atual aware

    if agora > DATA_FINAL:
        return render_template('encerramento.html'), 403 # Código HTTP 403 - Acesso proibido
    
    return render_template('index.html')

@app.route('/inscrever', methods=['POST'])
def inscrever():
    try:
        dados = {campo: request.form[campo] for campo in [
            'nome', 'rg', 'cpf', 'celular', 'email', 'empresa', 'bairro'
        ]}
        dados['autorizacao_dados'] = request.form.get('autorizacao_dados', 'Não')
        dados['autorizacao_imagem'] = request.form.get('autorizacao_imagem', 'Não')

        with sqlite3.connect('cadastros.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM inscritos WHERE cpf = ? OR rg = ? OR email = ?", (dados['cpf'], dados['rg'], dados['email']))
            if cursor.fetchone():
                return redirect(url_for('index', erro="CPF, RG ou E-mail já cadastrado!"))
            cursor.execute("INSERT INTO inscritos VALUES (NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?)", tuple(dados.values()))
            conn.commit()

        return redirect(url_for('success', mensagem="Cadastro realizado com sucesso!"))
    except KeyError as e:
        return f"Erro: Campo ausente no formulário -> {str(e)}", 400

@app.route('/success')
def success():
    mensagem = request.args.get('mensagem', "Inscrição concluída!")
    return render_template('success.html', mensagem=mensagem)

@app.route('/inscritos')
@login_required
def inscritos():
    with sqlite3.connect('cadastros.db') as conn:
        inscritos = conn.execute("SELECT * FROM inscritos").fetchall()

        # Define o número de inscritos por página
        inscritos_por_pagina = 20

        # Calcula o número total de páginas
        num_paginas = (len(inscritos) + inscritos_por_pagina - 1) // inscritos_por_pagina

        # Obtém o número da página atual (se não especificado, assume a primeira página)
        pagina_atual = int(request.args.get('pagina', 1))

        # Calcula o índice inicicial e final dos inscritos na página atual
        inicio = (pagina_atual - 1) * inscritos_por_pagina
        fim = min(pagina_atual * inscritos_por_pagina, len(inscritos))

        # Obtém os inscritos da página atual
        inscritos_pagina = inscritos[inicio:fim]

    return render_template('inscritos.html',
                           inscritos=inscritos_pagina,
                           pagina_atual=pagina_atual,
                           num_paginas=num_paginas)

@app.route('/remover/<int:id>', methods=['POST'])
@login_required
def remover_inscrito(id):
    with sqlite3.connect('cadastros.db') as conn:
        conn.execute("DELETE FROM inscritos WHERE id = ?", (id,))
        conn.commit()
    return redirect(url_for('inscritos'))

# Corta o texto caso ele exceda o limite da celula (será colado um limite na largura da célula até onde o texto pode ir)
def truncar_texto(pdf, texto, largura_maxima):
    #Guarda o texto original para comparar depois
    texto_original = texto

    # Trunca o texto para caber dentro da largura máxima da célula e adiciona '...' no final.
    while pdf.get_string_width(texto + "...") > largura_maxima and len (texto) > 0:
        texto = texto[:-1] # Remove um caractere por vez

    # Se o texto foi cortado, adiciona "..."
    return texto + "..." if texto != texto_original else texto

@app.route('/exportar_pdf_individual/<int:id>')
@login_required
def exportar_pdf_individual(id):
    with sqlite3.connect('cadastros.db') as conn:
        inscrito = conn.execute("SELECT * FROM inscritos WHERE id = ?", (id,)).fetchone()

    if inscrito is None:
        return "Inscrito não encontrado", 404

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    dados_inscrito = [
        ("RG:", inscrito[2], "CPF:", inscrito[3]),
        ("Celular:", inscrito[4], "Email:", inscrito[5]),
        ("Empresa:", inscrito[6], "Bairro:", inscrito[7]),
        ("Autorização de uso de dados:", "Sim" if inscrito[8] == "on" else "Não", "Autorização de uso da Imagem:", "Sim" if inscrito[9] == "on" else "Não"),
    ]

    pdf.cell(190, 10, f"Nome: {inscrito[1]}", 1, align='C')
    pdf.ln()

    for rotulo1, valor1, rotulo2, valor2 in dados_inscrito:
        largura_celula = 77 # Define a largura máxima da célula

        pdf.cell(95, 10, f"{rotulo1} {truncar_texto(pdf, valor1, largura_celula)}", 1, align='L')
        pdf.cell(95, 10, f"{rotulo2} {truncar_texto(pdf, valor2, largura_celula)}", 1, align='L')
        pdf.ln()

    pdf.cell(190, 9, "*A não comprovação da contribuição junto ao sindicato não dará direito ao prêmio!", 1, align='C')
    pdf.ln()

    # Gera um ID único para o nome do arquivo
    nome_arquivo = f"inscrito_{uuid.uuid4()}.pdf"
    caminho_pdf = os.path.join("static", nome_arquivo)
    pdf.output(caminho_pdf)

    return send_file(caminho_pdf, mimetype='application/pdf')

@app.route('/exportar_pdf/<int:pagina>')  # Nova rota com o número da página
@login_required
def exportar_pdf(pagina):
    with sqlite3.connect('cadastros.db') as conn:
        inscritos = conn.execute("SELECT * FROM inscritos").fetchall()

        # Define o número de inscritos por página
        inscritos_por_pagina = 20

        # Calcula o índice inicial e final dos inscritos na página atual
        inicio = (pagina - 1) * inscritos_por_pagina
        fim = min(pagina * inscritos_por_pagina, len(inscritos))

        # Obtém os inscritos da página atual
        inscritos_pagina = inscritos[inicio:fim]

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    dados_pdf = []
    for inscrito in inscritos_pagina:
        dados_inscrito = [
            ("RG:", inscrito[2], "CPF:", inscrito[3]),
            ("Celular:", inscrito[4], "Email:", inscrito[5]),
            ("Empresa:", inscrito[6], "Bairro:", inscrito[7]),
            ("Autorização de uso de dados:", "Sim" if inscrito[8] == "on" else "Não", "Autorização de uso da Imagem:", "Sim" if inscrito[9] == "on" else "Não"), # Condicional para "Sim" ou "Não"
        ]
        dados_pdf.append(dados_inscrito)

    for nome, dados_inscrito in zip([inscrito[1] for inscrito in inscritos_pagina], dados_pdf):
        # Título com nome (mesclado e centralizado)
        pdf.cell(190, 10, f"Nome: {nome}", 1, align='C')
        pdf.ln()

        for rotulo1, valor1, rotulo2, valor2 in dados_inscrito:
            largura_celula = 77 # Define a largura máxima da célula

            # Coluna A: Rótulo 1
            pdf.cell(95, 10, f"{rotulo1} {truncar_texto(pdf, valor1, largura_celula)}", 1, align='L')
            # Coluna B: Valor 1
            pdf.cell(95, 10, f"{rotulo2} {truncar_texto(pdf, valor2, largura_celula)}", 1, align='L')
            pdf.ln()

        #Linha de entendimento da contribuição junto ao sindicato
        pdf.cell(190, 9, "*A não comprovação da contribuição junto ao sindicato não dará direito ao prêmio!", 1, align='C')
        pdf.ln() #Quebra de linha

        pdf.ln()  # Espaço entre os registros

    caminho_pdf = f"static/inscritos_pagina_{pagina}.pdf"  # Nome do arquivo com o número da página
    os.makedirs("static", exist_ok=True)
    pdf.output(caminho_pdf)

    return redirect(f"/{caminho_pdf}")

@app.route('/exportar_planilha')
@login_required
def exportar_planilha():
    with sqlite3.connect('cadastros.db') as conn:
        inscritos = conn.execute("SELECT * FROM inscritos").fetchall()

    wb = Workbook() # Instancia a classe Workbook
    ws = wb.active # Acesse a planilha ativa

    # Adicione os cabeçalhos da planilha
    ws.append(['Nome', 'RG', 'CPF', 'Celular', 'Email', 'Empresa', 'Bairro', 'Autorização de uso de dados', 'Autorização de uso da imagem'])

    # Adicione o dados dos inscritos
    for inscrito in inscritos:
        ws.append([inscrito[1], inscrito[2], inscrito[3], inscrito[4], inscrito[5], inscrito[6], inscrito[7], "Sim" if inscrito[8] == "on" else "Não", "Sim" if inscrito[9] == "on" else "Não"])

    # Salve a planilha
    caminho_planilha = "static/inscritos.xlsx"
    wb.save(caminho_planilha)

    return redirect(f"/{caminho_planilha}")

if __name__ == '__main__':
    app.run(debug=True, port=7000)
